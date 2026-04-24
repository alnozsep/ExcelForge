import pytest


@pytest.fixture
def sample_pdf_bytes():
    # ダミーのPDFを作成 (PyMuPDFが必要)
    import fitz

    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 50), "Test Corporation")
    page2 = doc.new_page()
    page2.insert_text((50, 50), "DX Promotion Project")
    return doc.write()


@pytest.fixture
def empty_pdf_bytes():
    import fitz

    doc = fitz.open()
    doc.new_page()
    return doc.write()


@pytest.fixture
def corrupted_pdf_bytes():
    return b"%PDF-1.4\ncorrupted data"


@pytest.fixture
def sample_csv_bytes():
    return "会社名,担当者\n株式会社テスト,山田太郎\n".encode("utf-8")


@pytest.fixture
def sample_csv_sjis_bytes():
    return "会社名,担当者\n株式会社テスト,山田太郎\n".encode("shift_jis")


@pytest.fixture
def sample_xlsx_bytes():
    import openpyxl
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "株式会社テスト"
    ws["B1"] = 1500000

    # 結合セルテスト用
    ws.merge_cells("A3:C3")
    ws["A3"] = "補助金申請データ"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
