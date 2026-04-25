import openpyxl
import io

template_path = "/Users/nozakizai/Desktop/excelforge/test_data/invoice_template.xlsx"
wb = openpyxl.load_workbook(template_path)
found = False
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                print(f"Found placeholder: {cell.value} at {sheet}:{cell.coordinate}")
                found = True
if not found:
    print("No placeholders found.")
