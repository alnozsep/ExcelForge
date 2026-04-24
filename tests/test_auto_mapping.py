import io
import openpyxl
from app.core.excel_writer import write_to_template


def test_direct_coordinate_writing():
    # テンプレートの作成 (プレースホルダーなし)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name Label"
    ws["B1"] = ""  # ここに書き込みたい

    template_buffer = io.BytesIO()
    wb.save(template_buffer)
    template_bytes = template_buffer.getvalue()

    # AIからの抽出結果 (Sheet:Cell 形式)
    extracted_data = {
        "Sheet1:B1": "John Doe",
        "Sheet1:A1": "Modified Label",  # ラベルも上書きできるか
    }

    # 書き込み実行
    output_buffer = write_to_template(template_bytes, extracted_data)

    # 結果の検証
    result_wb = openpyxl.load_workbook(output_buffer)
    result_ws = result_wb["Sheet1"]
    assert result_ws["B1"].value == "John Doe"
    assert result_ws["A1"].value == "Modified Label"


def test_direct_coordinate_with_merged_cells():
    # テンプレートの作成 (結合セルあり)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataSheet"
    ws.merge_cells("B2:D3")

    template_buffer = io.BytesIO()
    wb.save(template_buffer)
    template_bytes = template_buffer.getvalue()

    # 結合範囲内のどこを指定してもマスターセルに書き込まれるべき
    extracted_data = {"DataSheet:C2": "Merged Value"}

    output_buffer = write_to_template(template_bytes, extracted_data)

    result_wb = openpyxl.load_workbook(output_buffer)
    result_ws = result_wb["DataSheet"]
    # B2 (マスターセル) に値が入っていることを確認
    assert result_ws["B2"].value == "Merged Value"
