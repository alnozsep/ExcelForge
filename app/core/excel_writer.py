"""
app/core/excel_writer.py

責務: Excelテンプレートをメモリ上で開き、抽出データを指定セルに書き込み、BytesIOとして返す。
"""

import io
import re
from typing import Optional, Any
import openpyxl

from app.models.enums import AppException, ErrorCode


def _is_merged_cell(worksheet, cell) -> bool:
    """セルが結合セルの一部かどうかを判定する"""
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False


def _get_merged_cell_master(worksheet, cell):
    """結合セルの左上（マスター）セルを返す"""
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 結合範囲の左上のセルを返す
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _has_formula(cell) -> bool:
    """セルに数式が含まれているかを判定する"""
    return cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    )


def write_to_template(
    template_bytes: bytes, extracted_data: dict, mapping_config: Optional[dict] = None
) -> io.BytesIO:
    """
    Excelテンプレートにデータを書き込む。
    """
    try:
        # メモリ上でExcelファイルを開く
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))

        # 1. マッピング設定に基づく書き込み
        if mapping_config and "mappings" in mapping_config:
            mappings = mapping_config.get("mappings", [])
            for mapping in mappings:
                key = mapping.get("key")
                sheet_name = (mapping.get("sheet") or "").strip()
                cell_ref = (mapping.get("cell") or "").strip()

                if not key or key not in extracted_data:
                    continue

                val = extracted_data[key]
                if val is None:
                    continue

                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active

                try:
                    target_cell = ws[cell_ref]
                    if _is_merged_cell(ws, target_cell):
                        target_cell = _get_merged_cell_master(ws, target_cell)
                    if not _has_formula(target_cell):
                        target_cell.value = val
                except Exception:
                    continue

        # 2. 直接セル指定（Sheet:Cell形式）の書き込み
        for key, val in extracted_data.items():
            if val is None or not isinstance(key, str) or ":" not in key:
                continue

            parts = key.split(":")
            if len(parts) == 2:
                sheet_name = parts[0].strip()
                cell_ref = parts[1].strip()
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    try:
                        target_cell = ws[cell_ref]
                        if _is_merged_cell(ws, target_cell):
                            target_cell = _get_merged_cell_master(ws, target_cell)
                        if not _has_formula(target_cell):
                            target_cell.value = val
                    except Exception:
                        continue

        # 3. プレースホルダー {{key}} を探して置換
        # mapping_configがない場合、または明示的に指定された場合の両方で動作させる
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        isinstance(cell.value, str)
                        and "{{" in cell.value
                        and "}}" in cell.value
                    ):
                        new_value: Any = cell.value
                        matches = re.findall(r"\{\{([^}]+)\}\}", cell.value)

                        has_replacement = False
                        for key in matches:
                            key_stripped = key.strip()
                            if (
                                key_stripped in extracted_data
                                and extracted_data[key_stripped] is not None
                            ):
                                val = extracted_data[key_stripped]
                                val_str = str(val)
                                new_value = new_value.replace(f"{{{{{key}}}}}", val_str)
                                has_replacement = True

                        if has_replacement:
                            # 完全に "{{key}}" のみで置換結果が数値の場合は数値として扱う
                            if (
                                len(matches) == 1
                                and cell.value.strip() == f"{{{{{matches[0]}}}}}"
                            ):
                                val = extracted_data[matches[0].strip()]
                                if isinstance(val, (int, float)):
                                    new_value = val

                            target_cell = cell
                            if _is_merged_cell(ws, target_cell):
                                target_cell = _get_merged_cell_master(ws, target_cell)

                            if not _has_formula(target_cell):
                                target_cell.value = new_value

        # BytesIOに保存して返す
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        raise AppException(
            ErrorCode.TEMPLATE_WRITE_ERROR, f"テンプレート書き込みエラー: {str(e)}"
        )

    except Exception as e:
        raise AppException(
            ErrorCode.TEMPLATE_WRITE_ERROR, f"テンプレート書き込みエラー: {str(e)}"
        )
