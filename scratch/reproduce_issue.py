import asyncio
import json
import os
from unittest.mock import AsyncMock, patch
from app.core.extractor import extract_data
from app.core.file_reader import read_file, FileType
from app.core.excel_writer import write_to_template

async def reproduce():
    # Load test data
    csv_path = "/Users/nozakizai/Desktop/excelforge/test_data/source_invoice_data.csv"
    template_path = "/Users/nozakizai/Desktop/excelforge/test_data/invoice_template.xlsx"
    
    with open(csv_path, "rb") as f:
        csv_bytes = f.read()
    with open(template_path, "rb") as f:
        template_bytes = f.read()
    
    source_text = csv_bytes.decode("utf-8")
    template_text = read_file(template_bytes, FileType.XLSX)
    
    print("--- Source Text ---")
    print(source_text)
    print("--- Template Text (Excerpt) ---")
    print("\n".join(template_text.split("\n")[:20]))
    
    # Mock Gemini to see what happens with a "bad" response or to simulate a "good" one
    # Let's see if we can identify any logic errors in the code without a real API call first.
    # But the user might be referring to a real API failure or a logic failure after API.
    
    # Let's simulate a likely response from Gemini for this data
    mock_response_data = {
        "請求書:A5": "グローバルイノベーション株式会社",
        "請求書:G3": "2026-04-30",
        "請求書:C7": 1045000,
        "請求書:B12": "AIコンサルティング費用(4月分)",
        "請求書:C12": 1,
        "請求書:E12": 500000,
        "請求書:F12": 500000,
        "請求書:B13": "クラウドインフラ構築支援",
        "請求書:C13": 1,
        "請求書:E13": 300000,
        "請求書:F13": 300000,
        "請求書:B14": "システム保守・ライセンス費",
        "請求書:C14": 1,
        "請求書:E14": 150000,
        "請求書:F14": 150000,
    }
    
    print("\n--- Simulating Write ---")
    output_buffer = write_to_template(template_bytes, mock_response_data)
    
    import openpyxl
    wb = openpyxl.load_workbook(output_buffer)
    ws = wb["請求書"]
    
    print(f"A5 (Customer): {ws['A5'].value}")
    print(f"G3 (Date): {ws['G3'].value}")
    print(f"C7 (Total): {ws['C7'].value}")
    print(f"B12 (Item 1): {ws['B12'].value}")
    print(f"F12 (Amount 1): {ws['F12'].value}")
    print(f"F13 (Amount 2): {ws['F13'].value}")
    print(f"F14 (Amount 3): {ws['F14'].value}")

if __name__ == "__main__":
    asyncio.run(reproduce())
