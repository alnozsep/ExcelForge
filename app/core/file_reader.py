"""
app/core/file_reader.py


責務: アップロードされたファイルをメモリ上で読み込み、
      テキスト文字列に変換して返す。


入力: ファイルバイナリ（bytes）、ファイル拡張子（str）
出力: 抽出されたテキスト（str）


制約:
  - ファイルシステムへの書き込みは行わない
  - すべてio.BytesIOを使用してメモリ上で処理する
  - 対応形式: PDF, CSV, XLSX, XLS
  - XLS形式はxlrdライブラリで対応する（openpyxlは非対応）
"""

import io
import re
from enum import Enum
import openpyxl
import xlrd
import fitz

from app.config import settings
from app.models.enums import FileReadError

# Zip爆弾対策用の安全制限値
MAX_TEXT_LENGTH: int = 100_000  # 最大10万文字
MAX_ESTIMATED_TOKENS: int = 50_000  # 最大5万トークン（概算）


class FileType(Enum):
    PDF = "pdf"
    CSV = "csv"
    XLSX = "xlsx"
    XLS = "xls"


def detect_file_type(filename: str) -> FileType:
    """
    ファイル名から拡張子を判定してFileTypeを返す。
    対応外の拡張子の場合はValueErrorを送出する。
    """
    if not filename:
        raise ValueError("Filename is empty")

    parts = filename.lower().split(".")
    if len(parts) < 2:
        raise ValueError(f"No extension found in filename: {filename}")

    ext = parts[-1]

    try:
        return FileType(ext)
    except ValueError:
        raise ValueError(f"Unsupported file extension: {ext}")


def validate_text_size(text: str, source_type: str = "") -> None:
    """
    抽出テキストのサイズを検証する。

    Zip爆弾対策: 1MBのPDFから10万文字以上のテキストが
    展開されるケースを防ぐ。
    """
    if len(text) > MAX_TEXT_LENGTH:
        raise FileReadError(
            f"抽出されたテキストが上限（{MAX_TEXT_LENGTH:,}文字）を"
            f"超えています（{len(text):,}文字）。"
            f"ページ数を減らしてお試しください。"
        )

    # トークン数の概算（日本語: 約1文字=1.5トークン）
    estimated_tokens = int(len(text) * 1.5)
    if estimated_tokens > MAX_ESTIMATED_TOKENS:
        raise FileReadError(
            "テキスト量がAI処理の上限を超えています。"
            "ファイルのページ数を減らしてお試しください。"
        )


def read_pdf(file_bytes: bytes) -> str:
    """PDF読み込み（ページ数制限付き）"""
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
    except Exception as e:
        raise FileReadError(f"PDFファイルを開けません: {type(e).__name__}")

    if len(doc) == 0:
        raise FileReadError("PDFファイルが空です。")

    text_parts = []
    for i in range(min(len(doc), settings.MAX_PDF_PAGES)):
        page = doc[i]
        page_text = page.get_text()
        text_parts.append(f"--- Page {i + 1} ---\n{page_text}")

    text = "\n".join(text_parts).strip()
    # ページヘッダのみ残って中身が空の場合の対応
    text_only = re.sub(r"--- Page \d+ ---", "", text).strip()
    if not text_only:
        raise FileReadError(
            "テキストが抽出できませんでした（スキャンPDFは非対応です）。"
        )

    validate_text_size(text, "pdf")
    return text


def read_file(file_bytes: bytes, file_type: FileType) -> str:
    """
    ファイルバイナリを受け取り、テキスト文字列に変換して返す。
    """
    if not file_bytes:
        raise FileReadError("ファイルデータが空です。")

    try:
        if file_type == FileType.PDF:
            return read_pdf(file_bytes)

        elif file_type == FileType.CSV:
            try:
                text = file_bytes.decode("utf-8")
            except UnicodeDecodeError:
                try:
                    text = file_bytes.decode("shift_jis")
                except Exception as e:
                    raise FileReadError(
                        f"CSVファイルの読み込みに失敗しました（エンコーディングエラー）: {type(e).__name__}"
                    )
            
            if not text.strip():
                raise FileReadError("CSVファイルが空です。")

            validate_text_size(text, "csv")
            return text

        elif file_type == FileType.XLSX:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            except Exception as e:
                raise FileReadError(
                    f"XLSXファイルの読み込みに失敗しました: {type(e).__name__}"
                )

            text_parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 有効なデータ範囲を取得
                min_row, max_row = ws.min_row, ws.max_row
                min_col, max_col = ws.min_column, ws.max_column
                
                # 範囲が広すぎる場合の制限（最大100行100列程度）
                max_row = min(max_row, 100)
                max_col = min(max_col, 26) # A-Z

                # 結合セルの範囲を把握
                merged_ranges = ws.merged_cells.ranges

                text_parts.append(f"Sheet: {sheet_name}")
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        # 結合セルかどうか判定
                        is_merged = False
                        master_coord = None
                        for m_range in merged_ranges:
                            if cell.coordinate in m_range:
                                is_merged = True
                                master_coord = openpyxl.utils.get_column_letter(m_range.min_col) + str(m_range.min_row)
                                break
                        
                        if is_merged and cell.coordinate != master_coord:
                            val = f"(part of merged cell {master_coord})"
                        else:
                            val = cell.value if cell.value is not None else "(empty)"
                        
                        text_parts.append(f"{cell.coordinate}: {val}")

            text = "\n".join(text_parts)
            validate_text_size(text, "xlsx")
            return text

        elif file_type == FileType.XLS:
            try:
                wb = xlrd.open_workbook(file_contents=file_bytes)
            except Exception as e:
                raise FileReadError(
                    f"XLSファイルの読み込みに失敗しました: {type(e).__name__}"
                )

            text_parts = []
            for sheet in wb.sheets():
                text_parts.append(f"Sheet: {sheet.name}")
                # XLSも100x26に制限
                nrows = min(sheet.nrows, 100)
                ncols = min(sheet.ncols, 26)
                
                for rowx in range(nrows):
                    for colx in range(ncols):
                        cell_value = sheet.cell_value(rowx, colx)
                        val = cell_value if cell_value != "" else "(empty)"
                        # coordinateをA1形式に変換
                        col_letter = xlrd.colname(colx)
                        coord = f"{col_letter}{rowx + 1}"
                        text_parts.append(f"{coord}: {val}")

            text = "\n".join(text_parts)
            validate_text_size(text, "xls")
            return text

        else:
            raise ValueError(f"未対応のファイル形式です: {file_type}")

    except FileReadError:
        raise
    except Exception as e:
        raise FileReadError(
            f"ファイルの読み込み中に予期せぬエラーが発生しました: {type(e).__name__}"
        )
