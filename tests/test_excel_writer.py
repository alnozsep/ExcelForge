import pytest
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from app.core.excel_writer import write_to_template


@pytest.fixture
def template_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    # プレースホルダー
    ws["A1"] = "会社名: {{company_name}}"

    # 結合セル
    ws.merge_cells("C1:D2")
    ws["C1"] = "{{address}}"

    # 数式セル
    ws["A2"] = "=SUM(1, 2)"

    # 書式設定
    font = Font(bold=True, color="FF0000")
    alignment = Alignment(horizontal="center")
    ws["B1"].font = font
    ws["B1"].alignment = alignment
    ws["B1"] = "{{amount}}"

    # 既存値（nullのテスト用）
    ws["A3"] = "既存データ"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_placeholder_replacement(template_bytes):
    """{{キー名}}プレースホルダーが正しく置換されること"""
    extracted_data = {"company_name": "株式会社テスト"}
    result_buf = write_to_template(template_bytes, extracted_data)

    wb = openpyxl.load_workbook(result_buf)
    ws = wb.active
    assert ws["A1"].value == "会社名: 株式会社テスト"


def test_mapping_config_cell_writing(template_bytes):
    """mapping_config指定のセルに正しく値が書き込まれること"""
    extracted_data = {"target": "ダイレクト入力"}
    mapping_config = {
        "mappings": [{"key": "target", "sheet": "Template", "cell": "E1"}]
    }

    result_buf = write_to_template(template_bytes, extracted_data, mapping_config)
    wb = openpyxl.load_workbook(result_buf)
    ws = wb.active
    assert ws["E1"].value == "ダイレクト入力"


def test_merged_cell_writing(template_bytes):
    """結合セルの左上に正しく書き込まれること"""
    extracted_data = {"address": "東京都新宿区"}
    result_buf = write_to_template(template_bytes, extracted_data)

    wb = openpyxl.load_workbook(result_buf)
    ws = wb.active
    assert ws["C1"].value == "東京都新宿区"
    # D1, C2, D2などはNoneまたは結合セルの状態を維持


def test_formula_cell_skipped(template_bytes):
    """数式セルへの書き込みがスキップされること"""
    extracted_data = {"formula": "上書きされるべきでない"}
    mapping_config = {
        "mappings": [{"key": "formula", "sheet": "Template", "cell": "A2"}]
    }

    result_buf = write_to_template(template_bytes, extracted_data, mapping_config)
    wb = openpyxl.load_workbook(result_buf)
    ws = wb.active
    # 数式セルは書き込みをスキップするため元の数式が維持されること
    assert str(ws["A2"].value).startswith("=")


def test_cell_format_preserved(template_bytes):
    """書き込み後もセルの書式（フォント・配置）が保持されること"""
    extracted_data = {"amount": 1000}
    result_buf = write_to_template(template_bytes, extracted_data)

    wb = openpyxl.load_workbook(result_buf)
    ws = wb.active
    assert ws["B1"].value == 1000
    assert ws["B1"].font.bold is True
    assert ws["B1"].font.color.rgb == "00FF0000"
    assert ws["B1"].alignment.horizontal == "center"


def test_null_value_preserves_existing(template_bytes):
    """null値の場合、既存セル値が消されないこと"""
    # プレースホルダーに対する値がない、またはmapping_configで指定した値がNone
    extracted_data = {"existing": None}
    mapping_config = {
        "mappings": [{"key": "existing", "sheet": "Template", "cell": "A3"}]
    }

    result_buf = write_to_template(template_bytes, extracted_data, mapping_config)
    wb = openpyxl.load_workbook(result_buf)
    ws = wb.active
    assert ws["A3"].value == "既存データ"


def test_output_is_bytesio(template_bytes):
    """戻り値がio.BytesIO型であること"""
    result_buf = write_to_template(template_bytes, {})
    assert isinstance(result_buf, io.BytesIO)


def test_output_is_valid_xlsx(template_bytes):
    """出力がopenpyxlで開ける有効なxlsxであること"""
    result_buf = write_to_template(template_bytes, {})
    try:
        openpyxl.load_workbook(result_buf)
    except Exception as e:
        pytest.fail(f"Invalid xlsx format: {e}")
